"""
发货模板填充功能
根据订单信息和配置文件，填充发货模板
支持三套映射关系（映射1、映射2、映射3）
支持从数据库或Excel文件加载配置
"""
import os
from typing import Callable, Optional, List, Dict, Any
from excel_toolkit.excel_lite import ExcelReader
from excel_toolkit.excel_lite import get_column_letter

# 导入openpyxl用于写入Excel文件
try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    openpyxl = None
    MergedCell = None

# 尝试导入数据库模块
try:
    from excel_toolkit.db_config import get_db_manager
    from excel_toolkit.db_operations import load_shipping_config_from_db, get_all_warehouses
    _DB_AVAILABLE = True
except ImportError:
    _DB_AVAILABLE = False
    get_db_manager = load_shipping_config_from_db = get_all_warehouses = None


def _safe_set_cell_value(sheet, row, col, value, logger=None):
    """
    安全地设置单元格的值，自动处理合并单元格
    
    Args:
        sheet: openpyxl 工作表对象
        row: 行号（1-based）
        col: 列号（1-based）
        value: 要设置的值
        logger: 日志输出函数（可选）
    
    Returns:
        True 如果成功设置，False 如果跳过（合并单元格）
    """
    try:
        cell = sheet.cell(row=row, column=col)
        
        # 检查是否为合并单元格（MergedCell可能为None如果openpyxl未正确导入）
        if MergedCell is not None and isinstance(cell, MergedCell):
            # 查找合并区域的主单元格（左上角）
            for merged_range in sheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and 
                    merged_range.min_col <= col <= merged_range.max_col):
                    # 在主单元格设置值
                    main_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    main_cell.value = value
                    if logger:
                        col_letter = get_column_letter(col)
                        logger(f"⚠️ 检测到合并单元格 {col_letter}{row}，已填充到主单元格")
                    return True
            # 如果没找到合并区域，跳过
            if logger:
                col_letter = get_column_letter(col)
                logger(f"⚠️ 跳过合并单元格 {col_letter}{row}")
            return False
        else:
            # 普通单元格，直接设置值
            cell.value = value
            return True
            
    except Exception as e:
        if logger:
            col_letter = get_column_letter(col)
            logger(f"❌ 设置单元格 {col_letter}{row} 失败: {e}")
        return False


def _normalize_symbols(text: str) -> str:
    """
    标准化中英文符号，统一转为英文符号
    
    Args:
        text: 待处理的文本
    
    Returns:
        符号标准化后的文本
    """
    if not text:
        return text
    # 中文符号 -> 英文符号
    symbol_map = {
        '（': '(', '）': ')',
        '【': '[', '】': ']',
        '｛': '{', '｝': '}',
        '，': ',', '。': '.',
        '：': ':', '；': ';',
        '"': '"', '"': '"',
        ''': "'", ''': "'",
        '－': '-', '—': '-',
        '／': '/',
    }
    for cn, en in symbol_map.items():
        text = text.replace(cn, en)
    return text


def _match_warehouse(value: str, warehouses: list, alias_map: dict = None) -> str:
    """
    匹配仓库名称，支持别名映射和符号标准化
    
    匹配规则（按优先级）:
    1. 完全匹配标准仓库代码
    2. 标准化符号后匹配
    3. 别名映射表查找
    4. 未匹配到返回原值
    
    Args:
        value: 待匹配的仓库名称
        warehouses: 标准仓库代码列表
        alias_map: 可选的别名映射字典
    
    Returns:
        匹配到的标准仓库代码，或原值
    """
    if not value or not warehouses:
        return value
    
    value = str(value).strip()
    
    # 1. 完全匹配（原始值）
    if value in warehouses:
        return value
    
    # 标准化符号后再匹配
    normalized_value = _normalize_symbols(value)
    
    # 2. 完全匹配（标准化后）
    if normalized_value in warehouses:
        return normalized_value
    
    # 3. 别名映射表查找（原始值和标准化值都尝试）
    if alias_map:
        if value in alias_map:
            return alias_map[value]
        if normalized_value in alias_map:
            return alias_map[normalized_value]
    
    # 未匹配到，返回原值
    return value


def _format_table(headers: list, rows: list, title: str = None) -> str:
    """
    格式化表格输出，用于日志美化
    
    Args:
        headers: 表头列表
        rows: 数据行列表
        title: 可选的表格标题
    
    Returns:
        格式化后的表格字符串
    """
    if not rows:
        return ""
    
    # 计算每列最大宽度
    col_widths = []
    for i, h in enumerate(headers):
        # 中文字符宽度计2
        max_w = sum(2 if ord(c) > 127 else 1 for c in str(h))
        for row in rows:
            if i < len(row):
                cell = str(row[i]) if row[i] is not None else ""
                w = sum(2 if ord(c) > 127 else 1 for c in cell)
                max_w = max(max_w, w)
        col_widths.append(max_w)
    
    # 填充单元格
    def pad_cell(text, width):
        text = str(text) if text is not None else ""
        text_w = sum(2 if ord(c) > 127 else 1 for c in text)
        return text + " " * (width - text_w)
    
    separator = "+" + "+".join("-" * (w + 2) for w in col_widths) + "+"
    
    lines = []
    if title:
        lines.append(f"\n┌─ {title}")
    lines.append(separator)
    
    # 表头
    header_line = "|" + "|".join(f" {pad_cell(h, col_widths[i])} " for i, h in enumerate(headers)) + "|"
    lines.append(header_line)
    lines.append(separator)
    
    # 数据行
    for row in rows:
        row_line = "|" + "|".join(f" {pad_cell(row[i] if i < len(row) else '', col_widths[i])} " for i in range(len(headers))) + "|"
        lines.append(row_line)
    
    lines.append(separator)
    return "\n".join(lines)


def load_config_mapping(config_file: str, mapping_choice: str = "映射1", logger: Callable = print, 
                       config_name: Optional[str] = None) -> Dict[str, Any]:
    """
    加载配置文件
    支持三套映射关系：子表1（映射1）、子表2（映射2）、子表3（映射3）
    如果数据库已启用，优先从数据库加载
    
    Args:
        config_file: 配置文件路径（如果数据库未启用时使用）
        mapping_choice: "映射1" 或 "映射2" 或 "映射3"，选择使用哪套映射关系
        logger: 日志输出函数
        config_name: 数据库中的配置名称（如果使用数据库）
    
    返回: {
        "column_mapping": {订单列名: 模板列名},
        "column_mapping_1": {订单列名: 模板列名},  # 映射1
        "column_mapping_2": {订单列名: 模板列名},  # 映射2
        "column_mapping_3": {订单列名: 模板列名},  # 映射3
        "warehouses": [仓库名称列表],
        "shipping_map": {仓库名: {承运商: 物流渠道}}
    }
    """
    # 尝试从数据库加载
    if _DB_AVAILABLE:
        db_manager = get_db_manager()
        if db_manager.config.is_enabled():
            try:
                db_config = load_shipping_config_from_db(config_name or "默认配置")
                if db_config:
                    logger("✓ 从数据库加载发货配置")
                    
                    # 根据选择确定使用的映射
                    if mapping_choice == "映射3" and db_config.get("column_mapping_3"):
                        db_config["column_mapping"] = db_config["column_mapping_3"]
                        logger(f"✅ 使用映射3")
                    elif mapping_choice == "映射2" and db_config.get("column_mapping_2"):
                        db_config["column_mapping"] = db_config["column_mapping_2"]
                        logger(f"✅ 使用映射2")
                    else:
                        db_config["column_mapping"] = db_config.get("column_mapping_1", {})
                        logger(f"✅ 使用映射1")
                    
                    return db_config
            except Exception as e:
                logger(f"⚠️ 从数据库加载失败: {e}，将尝试从文件加载")
    
    # 从Excel文件加载
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"配置文件不存在: {config_file}")
    
    wb = ExcelReader(config_file, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    
    if not sheet_names:
        wb.close()
        raise ValueError("配置文件没有任何工作表")
    
    result = {
        "column_mapping": {},  # 当前选择的映射
        "column_mapping_1": {},  # 映射1（子表1）
        "column_mapping_2": {},  # 映射2（子表2，如果存在）
        "column_mapping_3": {},  # 映射3（子表3，如果存在）
        "default_values_2": {},  # 映射2的默认填充值（第三列）
        "default_values_3": {},  # 映射3的默认填充值（第三列）
        "direct_fill_defaults_1": {},  # 映射1：A列为空时，直接填充到模板列的默认值 {模板列名: 默认值}
        "direct_fill_defaults_2": {},  # 映射2：A列为空时，直接填充到模板列的默认值
        "direct_fill_defaults_3": {},  # 映射3：A列为空时，直接填充到模板列的默认值
        "warehouses": [],
        "shipping_map": {},
        "case_conversion_rules": {}  # 大小写转换规则 {表头: {原值: 转换值}}
    }
    
    # 子表1是映射1
    if len(sheet_names) >= 1:
        mapping_sheet1 = wb[sheet_names[0]]
        
        for row in mapping_sheet1.iter_rows(min_row=2, values_only=True):
            if row[1]:  # B列必须有值
                template_col = str(row[1]).strip()
                
                # A列有值：正常的列映射
                if row[0]:
                    order_col = str(row[0]).strip()
                    result["column_mapping_1"][order_col] = template_col
                # A列为空：直接填充默认值到模板列
                else:
                    if len(row) >= 3 and row[2] is not None:
                        result["direct_fill_defaults_1"][template_col] = row[2]
        
        # 表格输出映射1
        if result["column_mapping_1"]:
            rows = [[k, v] for k, v in result["column_mapping_1"].items()]
            logger(_format_table(["订单列", "模板列"], rows, f"映射1（{sheet_names[0]}）"))
        
        # 输出直接填充默认值规则
        if result["direct_fill_defaults_1"]:
            rows = [[k, v] for k, v in result["direct_fill_defaults_1"].items()]
            logger(_format_table(["模板列", "默认值"], rows, f"映射1-直接填充（A列为空）"))
    
    # 子表2是映射2（如果存在）
    if len(sheet_names) >= 2:
        mapping_sheet2 = wb[sheet_names[1]]
        
        for row in mapping_sheet2.iter_rows(min_row=2, values_only=True):
            if row[1]:  # B列必须有值
                template_col = str(row[1]).strip()
                
                # A列有值：正常的列映射
                if row[0]:
                    order_col = str(row[0]).strip()
                    result["column_mapping_2"][order_col] = template_col
                    # 读取第三列作为默认填充值（如果存在）
                    if len(row) >= 3 and row[2] is not None:
                        result["default_values_2"][order_col] = row[2]
                # A列为空：直接填充默认值到模板列
                else:
                    if len(row) >= 3 and row[2] is not None:
                        result["direct_fill_defaults_2"][template_col] = row[2]
        
        # 表格输出映射2（包含默认值列）
        if result["column_mapping_2"]:
            rows = [[k, v, result["default_values_2"].get(k, "")] for k, v in result["column_mapping_2"].items()]
            logger(_format_table(["订单列", "模板列", "默认值"], rows, f"映射2（{sheet_names[1]}）"))
        
        # 输出直接填充默认值规则
        if result["direct_fill_defaults_2"]:
            rows = [[k, v] for k, v in result["direct_fill_defaults_2"].items()]
            logger(_format_table(["模板列", "默认值"], rows, f"映射2-直接填充（A列为空）"))
    
    # 子表3是映射3（如果存在）
    if len(sheet_names) >= 3:
        # 检查子表3是否是特殊sheet（仓库别名、大小写转换规则）
        sheet3_name = sheet_names[2].strip()
        if sheet3_name not in ["仓库别名", "大小写转换规则"]:
            mapping_sheet3 = wb[sheet_names[2]]
            
            for row in mapping_sheet3.iter_rows(min_row=2, values_only=True):
                if row[1]:  # B列必须有值
                    template_col = str(row[1]).strip()
                    
                    # A列有值：正常的列映射
                    if row[0]:
                        order_col = str(row[0]).strip()
                        result["column_mapping_3"][order_col] = template_col
                        # 读取第三列作为默认填充值（如果存在）
                        if len(row) >= 3 and row[2] is not None:
                            result["default_values_3"][order_col] = row[2]
                    # A列为空：直接填充默认值到模板列
                    else:
                        if len(row) >= 3 and row[2] is not None:
                            result["direct_fill_defaults_3"][template_col] = row[2]
            
            # 表格输出映射3（包含默认值列）
            if result["column_mapping_3"]:
                rows = [[k, v, result["default_values_3"].get(k, "")] for k, v in result["column_mapping_3"].items()]
                logger(_format_table(["订单列", "模板列", "默认值"], rows, f"映射3（{sheet_names[2]}）"))
            
            # 输出直接填充默认值规则
            if result["direct_fill_defaults_3"]:
                rows = [[k, v] for k, v in result["direct_fill_defaults_3"].items()]
                logger(_format_table(["模板列", "默认值"], rows, f"映射3-直接填充（A列为空）"))
    
    # 根据选择确定使用的映射
    if mapping_choice == "映射3" and result["column_mapping_3"]:
        result["column_mapping"] = result["column_mapping_3"]
        logger(f"✅ 使用映射3（子表3）")
    elif mapping_choice == "映射2" and result["column_mapping_2"]:
        result["column_mapping"] = result["column_mapping_2"]
        logger(f"✅ 使用映射2（子表2）")
    else:
        result["column_mapping"] = result["column_mapping_1"]
        logger(f"✅ 使用映射1（子表1）")
    
    # 从子表4开始是仓库的物流渠道映射（排除“仓库别名”和“大小写转换规则”sheet）
    for sheet_name in sheet_names[3:]:
        sheet_name_stripped = sheet_name.strip()
        if sheet_name_stripped in ["仓库别名", "大小写转换规则"]:
            continue
        warehouse_name = sheet_name_stripped
        result["warehouses"].append(warehouse_name)
        result["shipping_map"][warehouse_name] = {}
        
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                carrier = str(row[0]).strip()
                shipping_service = str(row[1]).strip()
                result["shipping_map"][warehouse_name][carrier] = shipping_service
        
    
    # 表格输出仓库物流渠道映射概要
    if result["warehouses"]:
        wh_rows = [[wh, len(result["shipping_map"].get(wh, {}))] for wh in result["warehouses"]]
        logger(_format_table(["仓库", "承运商数"], wh_rows, "仓库物流配置"))
    
    # 读取仓库别名映射（如果存在"仓库别名"sheet）
    result["warehouse_alias"] = {}
    if "仓库别名" in sheet_names:
        alias_sheet = wb["仓库别名"]
        for row in alias_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                alias = str(row[0]).strip()
                standard = str(row[1]).strip()
                result["warehouse_alias"][alias] = standard
        
        if result["warehouse_alias"]:
            alias_rows = [[k, v] for k, v in result["warehouse_alias"].items()]
            logger(_format_table(["别名", "标准代码"], alias_rows, "仓库别名映射"))
    
    # 读取大小写转换规则（如果存在"大小写转换规则"sheet）
    if "大小写转换规则" in sheet_names:
        conversion_sheet = wb["大小写转换规则"]
        for row in conversion_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                header = str(row[0]).strip()  # 生效表头
                original_value = str(row[1]).strip()  # 原值
                converted_value = str(row[2]).strip()  # 转换后的值
                
                if header not in result["case_conversion_rules"]:
                    result["case_conversion_rules"][header] = {}
                result["case_conversion_rules"][header][original_value] = converted_value
        
        if result["case_conversion_rules"]:
            # 统计每个表头的转换规则数量
            conv_rows = [[header, len(rules)] for header, rules in result["case_conversion_rules"].items()]
            logger(_format_table(["表头", "规则数"], conv_rows, "大小写转换规则"))
            # 详细输出每个表头的转换规则
            for header, rules in result["case_conversion_rules"].items():
                detail_rows = [[orig, conv] for orig, conv in rules.items()]
                logger(_format_table(["原值", "转换值"], detail_rows, f"  {header}"))
    
    wb.close()
    return result


def get_warehouses_from_config(config_file: str) -> List[str]:
    """
    从配置文件获取所有仓库名称
    子表1、子表2、子表3是映射关系，从子表4开始是仓库
    如果数据库已启用，优先从数据库获取
    """
    # 尝试从数据库获取
    if _DB_AVAILABLE:
        db_manager = get_db_manager()
        if db_manager.config.is_enabled():
            try:
                warehouses = get_all_warehouses()
                if warehouses:
                    return warehouses
            except Exception:
                pass
    
    # 从Excel文件获取
    if not os.path.exists(config_file):
        return []
    
    try:
        wb = ExcelReader(config_file, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        # 子表1、子表2、子表3是映射关系，从子表4开始是仓库（排除“仓库别名”和“大小写转换规则”）
        warehouses = [s for s in sheet_names[3:] if s.strip() not in ["仓库别名", "大小写转换规则"]]
        return warehouses
    except Exception:
        return []


def get_mapping_choices_from_config(config_file: str) -> List[str]:
    """
    从配置文件获取可用的映射关系选项
    返回: ["映射1", "映射2", "映射3"] 或 ["映射1", "映射2"] 或 ["映射1"]
    如果数据库已启用，优先从数据库获取
    """
    # 尝试从数据库获取
    if _DB_AVAILABLE:
        db_manager = get_db_manager()
        if db_manager.config.is_enabled():
            try:
                from excel_toolkit.db_operations import list_shipping_configs
                configs = list_shipping_configs("mapping1")
                mapping2_configs = list_shipping_configs("mapping2")
                
                choices = ["映射1"]
                if mapping2_configs:
                    choices.append("映射2")
                # 暂时假设数据库中也有映射3的逻辑，可以扩展
                # mapping3_configs = list_shipping_configs("mapping3")
                # if mapping3_configs:
                #     choices.append("映射3")
                return choices
            except Exception:
                pass
    
    # 从Excel文件获取
    if not os.path.exists(config_file):
        return ["映射1"]
    
    try:
        wb = ExcelReader(config_file, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        choices = ["映射1"]
        special_sheets = ["仓库别名", "大小写转换规则"]
        # 如果存在子表2且不是特殊sheet，添加映射2选项
        if len(sheet_names) >= 2:
            sheet2_name = sheet_names[1].strip()
            if sheet2_name not in special_sheets:
                choices.append("映射2")
        # 如果存在子表3且不是特殊sheet，添加映射3选项
        if len(sheet_names) >= 3:
            sheet3_name = sheet_names[2].strip()
            if sheet3_name not in special_sheets:
                choices.append("映射3")
        return choices
    except Exception:
        return ["映射1"]


def check_template_has_data(template_file: str, template_sheet_name: str) -> Dict[str, Any]:
    """
    检测模板文件是否已有数据
    
    Args:
        template_file: 模板文件路径
        template_sheet_name: 模板工作表名称
    
    Returns:
        {
            "has_data": bool,  # 是否有数据
            "data_rows": int,  # 数据行数（不含表头）
            "last_row": int    # 最后一行的行号
        }
    """
    result = {
        "has_data": False,
        "data_rows": 0,
        "last_row": 1  # 默认第1行是表头
    }
    
    try:
        # 使用openpyxl来检测，因为它能更准确地读取xlsx文件
        if _OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(template_file, read_only=True, data_only=True)
            if template_sheet_name not in wb.sheetnames:
                wb.close()
                return result
            
            sheet = wb[template_sheet_name]
            
            # 从第2行开始检查（第1行是表头）
            data_row_count = 0
            last_data_row = 1
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=10), start=2):
                # 检查该行是否有数据
                has_data_in_row = False
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        has_data_in_row = True
                        break
                
                if has_data_in_row:
                    data_row_count += 1
                    last_data_row = row_idx
            
            wb.close()
            
            result["has_data"] = data_row_count > 0
            result["data_rows"] = data_row_count
            result["last_row"] = last_data_row
        else:
            # 回退到ExcelReader
            wb = ExcelReader(template_file, read_only=True, data_only=True)
            if template_sheet_name not in wb.sheetnames:
                wb.close()
                return result
            
            sheet = wb[template_sheet_name]
            
            data_row_count = 0
            last_data_row = 1
            
            for row_idx in range(2, sheet.max_row + 1):
                has_data_in_row = False
                for col_idx in range(1, min(11, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data_in_row = True
                        break
                
                if has_data_in_row:
                    data_row_count += 1
                    last_data_row = row_idx
            
            wb.close()
            
            result["has_data"] = data_row_count > 0
            result["data_rows"] = data_row_count
            result["last_row"] = last_data_row
        
    except Exception:
        pass
    
    return result


def process_shipping_fill(
    order_file: str,
    order_sheet_name: str,
    template_file: str,
    template_sheet_name: str,
    config_file: str,
    logger: Callable = print,
    warehouse_filter: Optional[List[str]] = None,
    mapping_choice: str = "映射1",
    fill_mode: str = "overwrite"
) -> str:
    """
    执行发货模板填充
    
    Args:
        order_file: 订单信息Excel文件路径
        order_sheet_name: 订单信息工作表名称
        template_file: 发货模板Excel文件路径
        template_sheet_name: 发货模板工作表名称
        config_file: 配置文件路径
        logger: 日志输出函数
        warehouse_filter: 要填充的仓库列表（None或空列表表示全部）
        mapping_choice: 选择使用的映射关系（"映射1" 或 "映射2" 或 "映射3"）
        fill_mode: 填充模式（"overwrite"=覆盖模式，从第2行开始；"append"=追加模式，在现有数据后追加）
    
    Returns:
        处理结果消息
    """
    # 1. 加载配置
    logger("=" * 50)
    logger("开始填充发货模板...")
    logger(f"配置文件: {config_file}")
    logger(f"使用映射关系: {mapping_choice}")
    
    config = load_config_mapping(config_file, mapping_choice, logger)
    column_mapping = config["column_mapping"]
    shipping_map = config["shipping_map"]
    warehouse_alias = config.get("warehouse_alias", {})
    default_values = config.get("default_values_2", {}) if mapping_choice == "映射2" else config.get("default_values_3", {}) if mapping_choice == "映射3" else {}
    
    # 获取直接填充默认值规则（A列为空的情况）
    if mapping_choice == "映射3":
        direct_fill_defaults = config.get("direct_fill_defaults_3", {})
    elif mapping_choice == "映射2":
        direct_fill_defaults = config.get("direct_fill_defaults_2", {})
    else:
        direct_fill_defaults = config.get("direct_fill_defaults_1", {})
    
    if not column_mapping:
        raise ValueError("配置文件中没有找到列映射关系")
    
    # 2. 打开订单文件
    logger(f"打开订单文件: {order_file}")
    order_wb = ExcelReader(order_file, read_only=True, data_only=True)
    template_wb = None
    
    try:
        if order_sheet_name not in order_wb.sheetnames:
            raise ValueError(f"订单文件中不存在工作表: {order_sheet_name}")
        order_sheet = order_wb[order_sheet_name]
        
        # 3. 打开模板文件（可写模式）- 使用openpyxl因为需要写入
        logger(f"打开模板文件: {template_file}")
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("需要openpyxl库来写入Excel文件，请安装: pip install openpyxl")
        template_wb = openpyxl.load_workbook(template_file)
        if template_sheet_name not in template_wb.sheetnames:
            raise ValueError(f"模板文件中不存在工作表: {template_sheet_name}")
        template_sheet = template_wb[template_sheet_name]
    
        # 4. 构建订单表头映射 {列名: 列索引}
        order_header_to_col = {}
        for col_idx, cell in enumerate(order_sheet[1], start=1):
            if cell.value:
                order_header_to_col[str(cell.value).strip()] = col_idx
        # 表格输出订单表头
        header_rows = [[i+1, name] for i, name in enumerate(order_header_to_col.keys())]
        logger(_format_table(["#", "订单列名"], header_rows, f"订单表头 ({len(order_header_to_col)}列)"))
        
        # 5. 构建模板表头映射 {列名: 列索引}
        template_header_to_col = {}
        for col_idx, cell in enumerate(template_sheet[1], start=1):
            if cell.value:
                template_header_to_col[str(cell.value).strip()] = col_idx
        logger(f"模板表头: {len(template_header_to_col)} 列")
        
        # 6. 找到关键列（智能查找，支持不同列名）
        # 6.1 查找仓库列（在订单文件中）- 优先从映射关系中查找
        warehouse_order_col = None
        warehouse_order_col_name = None
        warehouse_template_col_name = None
        
        # 方法一: 从映射关系中查找映射到"Warehouse Code/仓库代码"的源列
        for order_col, template_col in column_mapping.items():
            if "Warehouse" in template_col or "仓库" in template_col:
                warehouse_template_col_name = template_col
                warehouse_order_col_name = order_col
                warehouse_order_col = order_header_to_col.get(order_col)
                logger(f"✓ 从映射关系找到仓库列: {order_col} -> {template_col}")
                break
        
        # 方法二: 如果映射关系中没找到，尝试用常见列名匹配
        if not warehouse_order_col:
            for possible_name in ["仓库", "发货仓", "仓库代码", "Warehouse", "Warehouse Code"]:
                if possible_name in order_header_to_col:
                    warehouse_order_col = order_header_to_col[possible_name]
                    warehouse_order_col_name = possible_name
                    logger(f"✓ 通过常见列名找到仓库列: {possible_name}")
                    break
        
        if not warehouse_order_col:
            logger("⚠️ 警告: 订单文件中未找到'仓库'相关列，将跳过仓库筛选")
        
        # 6.2 确保模板中的仓库列名已设置
        if not warehouse_template_col_name:
            warehouse_template_col_name = "Warehouse Code/仓库代码"
        
        # 承运商列：优先从映射关系中查找
        carrier_template_col_name = None
        carrier_order_col_name = None
        # 尝试在映射关系中查找包含"承运商"或"carrier"的键
        for order_col, template_col in column_mapping.items():
            if "承运商" in order_col or "carrier" in order_col.lower():
                carrier_template_col_name = template_col
                carrier_order_col_name = order_col
                break
        if not carrier_template_col_name:
            carrier_template_col_name = "Carrier/承运商"
        
        # 物流渠道列（固定名称）
        shipping_service_col_name = "Shipping Service/物流渠道"
        
        # 州列：优先从映射关系中查找
        state_template_col_name = None
        for order_col, template_col in column_mapping.items():
            if "省份" in order_col or "州" in order_col or "state" in order_col.lower():
                state_template_col_name = template_col
                break
        if not state_template_col_name:
            state_template_col_name = "Recipient State/省/州"
        
        # 获取模板中的列索引
        warehouse_template_col = template_header_to_col.get(warehouse_template_col_name)
        carrier_template_col = template_header_to_col.get(carrier_template_col_name)
        shipping_service_col = template_header_to_col.get(shipping_service_col_name)
        state_template_col = template_header_to_col.get(state_template_col_name)
        
        # 表格输出关键列映射
        key_cols = [
            ["仓库", warehouse_order_col_name or "-", warehouse_template_col_name, warehouse_template_col or "-"],
            ["承运商", carrier_order_col_name or "-", carrier_template_col_name, carrier_template_col or "-"],
            ["物流渠道", "-", shipping_service_col_name, shipping_service_col or "-"],
            ["州/省份", "-", state_template_col_name, state_template_col or "-"],
        ]
        logger(_format_table(["字段", "订单列", "模板列", "索引"], key_cols, "关键列映射"))
        
        # 7. 确定填充起始行
        if fill_mode == "append":
            # 追加模式：检测现有数据的最后一行
            data_check = check_template_has_data(template_file, template_sheet_name)
            if data_check["has_data"]:
                template_row = data_check["last_row"] + 1
                logger(f"📋 追加模式: 检测到 {data_check['data_rows']} 行现有数据，从第 {template_row} 行开始填充")
            else:
                template_row = 2
                logger(f"📋 追加模式: 模板无数据，从第 2 行开始填充")
        else:
            # 覆盖模式：从第2行开始
            template_row = 2
            logger(f"📋 覆盖模式: 从第 2 行开始填充（将覆盖现有数据）")
        
        # 8. 开始填充数据
        filled_rows = 0
        skipped_rows = 0
        shipping_filled = 0
        state_converted = 0
        default_filled = 0  # 默认值填充计数
        direct_filled_rows = 0  # 直接填充行数（A列为空）
        country_converted = 0  # 国家名称转换计数
        case_converted = 0  # 大小写转换计数
        errors = []
        
        
        for order_row_idx in range(2, order_sheet.max_row + 1):
            # 获取订单中的仓库值
            order_warehouse_value = None
            if warehouse_order_col:
                cell_value = order_sheet.cell(row=order_row_idx, column=warehouse_order_col).value
                if cell_value:
                    order_warehouse_value = str(cell_value).strip()
            
            # 匹配仓库名称（支持别名映射）
            matched_warehouse = None
            if order_warehouse_value:
                all_warehouses = list(shipping_map.keys())
                matched_warehouse = _match_warehouse(order_warehouse_value, all_warehouses, warehouse_alias)
            
            # 应用仓库筛选
            if warehouse_filter:
                # 如果启用了仓库筛选，无仓库信息的行也跳过
                if not order_warehouse_value:
                    skipped_rows += 1
                    continue
                # 用匹配后的仓库名进行筛选
                if matched_warehouse not in warehouse_filter:
                    skipped_rows += 1
                    continue
            
            # 检查是否为空行（跳过）
            first_cell = order_sheet.cell(row=order_row_idx, column=1).value
            if first_cell is None:
                # 检查整行是否都为空
                all_empty = True
                for col_idx in range(1, min(10, order_sheet.max_column + 1)):
                    if order_sheet.cell(row=order_row_idx, column=col_idx).value:
                        all_empty = False
                        break
                if all_empty:
                    continue
            
            # 7.1 根据列映射填充数据
            for order_col_name, template_col_name in column_mapping.items():
                order_col = order_header_to_col.get(order_col_name)
                template_col = template_header_to_col.get(template_col_name)
                
                if order_col and template_col:
                    value = order_sheet.cell(row=order_row_idx, column=order_col).value
                    
                    # 映射2空值默认填充：如果订单列为空且有默认值，则使用默认值
                    if (value is None or str(value).strip() == "") and order_col_name in default_values:
                        value = default_values[order_col_name]
                        default_filled += 1
                    
                    # 州名转换（如果是省份列）
                    if order_col_name == "省份" or "省份" in order_col_name or "state" in order_col_name.lower():
                        if value:
                            value_str = str(value).strip()
                            # 只有当不是2个字符的缩写时才转换
                            if len(value_str) != 2:
                                try:
                                    from excel_toolkit.states import get_state_abbreviation
                                    abbr = get_state_abbreviation(value_str)
                                    if abbr:
                                        value = abbr
                                        state_converted += 1
                                except Exception:
                                    pass
                    
                    # 国家全称转简称（如果是国家列）
                    if "国家" in order_col_name or "country" in order_col_name.lower() or "Country" in template_col_name:
                        if value:
                            value_str = str(value).strip()
                            # 只有当不是2个字符的简称时才转换
                            if len(value_str) != 2:
                                country_abbr_map = {
                                    "Canada": "CA",
                                    "canada": "CA",
                                    "CANADA": "CA",
                                    "加拿大": "CA",
                                    "United States": "US",
                                    "united states": "US",
                                    "UNITED STATES": "US",
                                    "United States of America": "US",
                                    "USA": "US",
                                    "美国": "US",
                                }
                                if value_str in country_abbr_map:
                                    value = country_abbr_map[value_str]
                                    country_converted += 1
                    
                    # 应用配置文件中的大小写转换规则
                    case_conversion_rules = config.get("case_conversion_rules", {})
                    if case_conversion_rules and template_col_name in case_conversion_rules:
                        if value:
                            value_str = str(value).strip()
                            rules = case_conversion_rules[template_col_name]
                            # 尝试精确匹配
                            if value_str in rules:
                                value = rules[value_str]
                                case_converted += 1
                            # 尝试不区分大小写匹配
                            else:
                                for orig, conv in rules.items():
                                    if value_str.upper() == orig.upper():
                                        value = conv
                                        case_converted += 1
                                        break
                    
                    # 仓库列：将别名转换为标准代码
                    if "Warehouse" in template_col_name or "仓库" in template_col_name:
                        if value:
                            value_str = str(value).strip()
                            # 查找是否有别名映射
                            if value_str in warehouse_alias:
                                value = warehouse_alias[value_str]
                    
                    # 安全地设置单元格值（自动处理合并单元格）
                    _safe_set_cell_value(template_sheet, template_row, template_col, value)
            
            # 7.2 直接填充默认值（A列为空的情况）
            row_had_direct_fill = False
            if direct_fill_defaults:
                for template_col_name, default_value in direct_fill_defaults.items():
                    template_col = template_header_to_col.get(template_col_name)
                    if template_col:
                        # 安全地设置单元格值（自动处理合并单元格）
                        if _safe_set_cell_value(template_sheet, template_row, template_col, default_value):
                            row_had_direct_fill = True
            if row_had_direct_fill:
                direct_filled_rows += 1
            
            # 7.3 填充物流渠道
            if shipping_service_col and warehouse_template_col and carrier_template_col:
                warehouse = template_sheet.cell(row=template_row, column=warehouse_template_col).value
                carrier = template_sheet.cell(row=template_row, column=carrier_template_col).value
                
                if warehouse and carrier:
                    warehouse_str = str(warehouse).strip()
                    carrier_str = str(carrier).strip()
                    
                    # 匹配仓库名称（支持别名映射）
                    matched_wh = _match_warehouse(warehouse_str, list(shipping_map.keys()), warehouse_alias)
                    
                    # 查找物流渠道
                    if matched_wh in shipping_map:
                        wh_map = shipping_map[matched_wh]
                        if carrier_str in wh_map:
                            shipping_service = wh_map[carrier_str]
                            # 安全地设置单元格值（自动处理合并单元格）
                            if _safe_set_cell_value(template_sheet, template_row, shipping_service_col, shipping_service):
                                shipping_filled += 1
                        else:
                            errors.append(f"行{template_row}: 仓库[{matched_wh}]未找到承运商[{carrier_str}]的映射")
                    else:
                        errors.append(f"行{template_row}: 未找到仓库[{warehouse_str}]的配置")
            
            filled_rows += 1
            template_row += 1
        
        # 8. 保存结果（尝试直接覆盖，失败则保存到备份文件）
        save_success = False
        saved_path = template_file
        
        try:
            template_wb.save(template_file)
            save_success = True
        except PermissionError:
            # 文件被占用，尝试保存到备份文件
            logger("⚠️ 模板文件被占用（可能在Excel中打开），尝试保存到备份文件...")
            
            # 生成备份文件名
            base_name, ext = os.path.splitext(template_file)
            backup_path = f"{base_name}_已填充{ext}"
            
            # 如果备份文件也存在，添加数字后缀
            counter = 1
            while os.path.exists(backup_path):
                backup_path = f"{base_name}_已填充_{counter}{ext}"
                counter += 1
            
            try:
                template_wb.save(backup_path)
                saved_path = backup_path
                save_success = True
                logger(f"✅ 已保存到备份文件: {backup_path}")
            except Exception as e2:
                raise PermissionError(f"无法保存文件。请关闭Excel中打开的模板文件后重试。\n原始错误: {e2}")
    
    finally:
        # 9. 清理 - 确保文件正确关闭
        order_wb.close()
        if template_wb:
            template_wb.close()
    
    # 10. 输出统计
    logger("=" * 50)
    logger(f"✅ 填充完成！")
    if saved_path != template_file:
        logger(f"   - 保存位置: {os.path.basename(saved_path)}")
    logger(f"   - 填充行数: {filled_rows}")
    if skipped_rows > 0:
        logger(f"   - 跳过行数: {skipped_rows} (仓库筛选)")
    logger(f"   - 物流渠道填充: {shipping_filled}")
    if state_converted > 0:
        logger(f"   - 州名转换: {state_converted}")
    if default_filled > 0:
        logger(f"   - 默认值填充: {default_filled}")
    if direct_filled_rows > 0:
        logger(f"   - 直接填充: {direct_filled_rows} 行 (A列为空)")
    if country_converted > 0:
        logger(f"   - 国家名转换: {country_converted}")
    if case_converted > 0:
        logger(f"   - 大小写转换: {case_converted}")
    
    if errors:
        logger(f"⚠️ 警告 ({len(errors)} 条):")
        for err in errors[:10]:  # 最多显示10条
            logger(f"   {err}")
        if len(errors) > 10:
            logger(f"   ... 还有 {len(errors) - 10} 条警告")
    
    result_msg = f"填充完成！共填充 {filled_rows} 行"
    if skipped_rows > 0:
        result_msg += f"，跳过 {skipped_rows} 行"
    if shipping_filled > 0:
        result_msg += f"，物流渠道 {shipping_filled} 行"
    if state_converted > 0:
        result_msg += f"，州名转换 {state_converted} 次"
    if default_filled > 0:
        result_msg += f"，默认值填充 {default_filled} 次"
    if direct_filled_rows > 0:
        result_msg += f"，直接填充 {direct_filled_rows} 行"
    if country_converted > 0:
        result_msg += f"，国家名转换 {country_converted} 次"
    if case_converted > 0:
        result_msg += f"，大小写转换 {case_converted} 次"
    
    return result_msg

