from excel_toolkit.excel_lite import ExcelReader, ExcelWriter
from excel_toolkit.excel_lite import column_index_from_string
import os
from typing import Dict, List, Optional, Tuple, Any, Callable

# 导入openpyxl用于写入Excel文件
try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    openpyxl = None

def identify_header_mapping(headers: List[Any]) -> Dict[str, Optional[str]]:
    """
    根据表头列表猜测列映射
    返回: {'sku': 'SKU', 'l': '长', 'w': '宽', 'h': '高', 'wt': '重量'}
    """
    mapping = {'sku': None, 'l': None, 'w': None, 'h': None, 'wt': None}
    
    # 关键词列表 (小写)
    keywords = {
        'sku': ['sku', 'item', 'part', '编码', '料号', '产品代码'],
        'l': ['length', 'long', 'len', 'l', '长', '长度'],
        'w': ['width', 'wide', 'wid', 'w', '宽', '宽度'],
        'h': ['height', 'high', 'hgt', 'h', '高', '高度'],
        'wt': ['weight', 'weigh', 'wgt', 'wt', 'kg', 'g', '重', '重量', '单件重量']
    }
    
    # 预处理表头，减少循环中的重复操作
    headers_lower = [(h, str(h).lower().strip()) for h in headers if h]
    
    for key, kws in keywords.items():
        for h_orig, h_lower in headers_lower:
            if mapping[key]: break # 找到一个就停止
            if any(kw in h_lower for kw in kws):
                mapping[key] = h_orig
    
    return mapping

def parse_sku_bundle(sku_string: Any, logger: Optional[Callable] = None) -> Optional[List[Tuple[str, int]]]:
    sku_string = str(sku_string)
    if '+' not in sku_string and '*' not in sku_string:
        return None
    
    # 优化分割逻辑
    parts = sku_string.split('+') if '+' in sku_string else [sku_string]
    
    bundle = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if '*' in part:
            sub_parts = part.split('*')
            sku = sub_parts[0].strip()
            try:
                qty = int(sub_parts[1].strip())
                bundle.append((sku, qty))
            except (ValueError, IndexError):
                msg = f"  > 警告：SKU组合 '{part}' 格式错误，已跳过。"
                if logger: logger(msg)
        else:
            bundle.append((part.strip(), 1))
    return bundle


def process_skus(
    file_name: str, 
    sku_db_file: str, 
    db_col_map: Optional[Dict[str, str]] = None, 
    target_col_map: Optional[Dict[str, str]] = None, 
    logger: Callable[[str], None] = print,
    db_sheet_name: Optional[str] = None,
    ignore_qty: bool = False,
    order_sheet_name: Optional[str] = None
) -> Dict[str, int]:
    """
    智能填充SKU信息（支持外部数据库和灵活列映射）
    
    Args:
        db_sheet_name: SKU数据库工作表名称，如果为None则自动选择
        ignore_qty: 是否忽略数量列，按单个SKU计算
        order_sheet_name: 订单文件工作表名称，如果为None则处理所有工作表
    
    Returns:
        Dict with keys: 'sheets_processed', 'rows_filled'
        
    Raises:
        FileNotFoundError, ValueError, Exception
    """
    DEFAULT_DB_SHEET_NAME = "商品资料"
    # 默认映射
    if not db_col_map:
        db_col_map = {'sku': 'SKU', 'l': '长', 'w': '宽', 'h': '高', 'wt': '单件重量'}
    
    if not os.path.exists(sku_db_file):
        raise FileNotFoundError(f"SKU数据库文件不存在: {sku_db_file}")
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"订单文件不存在: {file_name}")

    # 1. 加载SKU数据库文件
    file_ext = os.path.splitext(sku_db_file)[1].lower()
    sku_database = {}
    
    if file_ext in ['.xlsx', '.xlsm']:
        # 使用openpyxl读取新版Excel
        try:
            logger(f"正在加载SKU数据库: {sku_db_file}...")
            db_wb = ExcelReader(sku_db_file, read_only=True, data_only=True)
        except Exception as e:
            raise Exception(f"加载SKU数据库失败: {e}")
        
        try:
            # 使用指定的工作表或自动选择
            if db_sheet_name and db_sheet_name in db_wb.sheetnames:
                data_ws = db_wb[db_sheet_name]
                logger(f"使用指定工作表: '{db_sheet_name}'")
            elif DEFAULT_DB_SHEET_NAME in db_wb.sheetnames:
                data_ws = db_wb[DEFAULT_DB_SHEET_NAME]
                logger(f"使用工作表: '{DEFAULT_DB_SHEET_NAME}'")
            else:
                data_ws = db_wb.worksheets[0]
                logger(f"未找到指定工作表，默认使用第一个工作表: '{data_ws.title}'")
                
            # 获取数据库表头映射到列索引
            headers = []
            for row in data_ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = list(row)
                break
                
            if not headers:
                 raise ValueError("SKU数据库似乎是空的（未找到表头）")

            db_headers_idx = {}
            for key, header_name in db_col_map.items():
                if header_name in headers:
                    db_headers_idx[key] = headers.index(header_name)
                else:
                    pass

            required_keys = ['sku', 'l', 'w', 'h', 'wt']
            missing_keys = [k for k in required_keys if k not in db_headers_idx]
            if missing_keys:
                missing_names = [db_col_map.get(k, k) for k in missing_keys]
                raise ValueError(f"SKU数据库缺少以下必需列: {', '.join(missing_names)}")
            
            # 2. 构建SKU数据库（内存字典）
            logger("正在构建SKU数据库...")
            
            idx_sku = db_headers_idx['sku']
            idx_l = db_headers_idx['l']
            idx_w = db_headers_idx['w']
            idx_h = db_headers_idx['h']
            idx_wt = db_headers_idx['wt']
            
            for row in data_ws.iter_rows(min_row=2, values_only=True):
                try:
                    sku_val = row[idx_sku]
                    if not sku_val: continue
                    
                    def to_float(v):
                        try: return float(v)
                        except: return 0.0
                    
                    l = to_float(row[idx_l])
                    w = to_float(row[idx_w])
                    h = to_float(row[idx_h])
                    wt = to_float(row[idx_wt])
                    
                    sku_database[sku_val] = {
                        "length": l, "width": w, "height": h, "weight": wt,
                        "volume": l * w * h
                    }
                except IndexError:
                    continue
            
            logger(f"SKU数据库构建完成，共加载 {len(sku_database)} 个SKU。")
            
        finally:
            db_wb.close()
            
    elif file_ext == '.xls':
        # 使用xlrd读取旧版Excel
        try:
            import xlrd
            logger(f"正在加载SKU数据库: {sku_db_file}...")
            db_wb = xlrd.open_workbook(sku_db_file)
        except Exception as e:
            raise Exception(f"加载SKU数据库失败: {e}")
        
        # 使用指定的工作表或自动选择
        sheet_names = db_wb.sheetnames
        if db_sheet_name and db_sheet_name in sheet_names:
            data_ws = db_wb.sheet_by_name(db_sheet_name)
            logger(f"使用指定工作表: '{db_sheet_name}'")
        elif DEFAULT_DB_SHEET_NAME in sheet_names:
            data_ws = db_wb.sheet_by_name(DEFAULT_DB_SHEET_NAME)
            logger(f"使用工作表: '{DEFAULT_DB_SHEET_NAME}'")
        else:
            data_ws = db_wb.sheet_by_index(0)
            logger(f"未找到指定工作表，默认使用第一个工作表: '{data_ws.name}'")
        
        if data_ws.nrows == 0:
            raise ValueError("SKU数据库似乎是空的（未找到表头）")
        
        # 获取表头
        headers = [cell.value for cell in data_ws.row(0)]
        
        db_headers_idx = {}
        for key, header_name in db_col_map.items():
            if header_name in headers:
                db_headers_idx[key] = headers.index(header_name)
            else:
                pass

        required_keys = ['sku', 'l', 'w', 'h', 'wt']
        missing_keys = [k for k in required_keys if k not in db_headers_idx]
        if missing_keys:
            missing_names = [db_col_map.get(k, k) for k in missing_keys]
            raise ValueError(f"SKU数据库缺少以下必需列: {', '.join(missing_names)}")
        
        # 构建SKU数据库
        logger("正在构建SKU数据库...")
        
        idx_sku = db_headers_idx['sku']
        idx_l = db_headers_idx['l']
        idx_w = db_headers_idx['w']
        idx_h = db_headers_idx['h']
        idx_wt = db_headers_idx['wt']
        
        for row_idx in range(1, data_ws.nrows):
            try:
                row = data_ws.row(row_idx)
                sku_val = row[idx_sku].value
                if not sku_val: continue
                
                def to_float(v):
                    try: return float(v)
                    except: return 0.0
                
                l = to_float(row[idx_l].value)
                w = to_float(row[idx_w].value)
                h = to_float(row[idx_h].value)
                wt = to_float(row[idx_wt].value)
                
                sku_database[sku_val] = {
                    "length": l, "width": w, "height": h, "weight": wt,
                    "volume": l * w * h
                }
            except IndexError:
                continue
        
        logger(f"SKU数据库构建完成，共加载 {len(sku_database)} 个SKU。")
    else:
        raise ValueError(f"不支持的文件格式: {file_ext}，请使用 .xlsx, .xlsm 或 .xls 文件")
    
    # 3. 加载订单文件
    order_file_ext = os.path.splitext(file_name)[1].lower()
    is_xls_order = (order_file_ext == '.xls')
    output_file_name = file_name
    
    if is_xls_order:
        logger(f"正在加载订单文件: {file_name}...")
        logger("⚠️  检测到.xls格式，将直接处理并输出为.xlsx格式")
        
        # 直接使用xlrd读取，处理后保存为xlsx
        import xlrd
        try:
            xls_wb = xlrd.open_workbook(file_name, formatting_info=False)
            logger(f"  找到 {len(xls_wb.sheetnames)} 个工作表")
        except Exception as e:
            raise Exception(f"加载.xls文件失败: {e}")
        
        # 输出文件名改为.xlsx
        output_file_name = file_name.rsplit('.', 1)[0] + '_processed.xlsx'
        logger(f"  处理后将保存为: {os.path.basename(output_file_name)}")
        
        # 转换为openpyxl工作簿以便编辑
        logger("  正在转换数据格式...")
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        for sheet_idx, sheet_name in enumerate(xls_wb.sheetnames):
            xls_sheet = xls_wb.sheet_by_index(sheet_idx)
            logger(f"    转换工作表 '{sheet_name}' ({xls_sheet.nrows} 行)...")
            
            xlsx_sheet = wb.create_sheet(title=sheet_name)
            
            # 批量转换数据
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    value = xls_sheet.cell_value(row_idx, col_idx)
                    xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
        
        logger("  ✅ 数据转换完成")
        
    else:
        # 加载.xlsx/.xlsm文件 - 使用openpyxl因为需要写入
        try:
            logger(f"正在加载订单文件: {file_name}...")
            if not _OPENPYXL_AVAILABLE:
                raise ImportError("需要openpyxl库来处理Excel文件，请安装: pip install openpyxl")
            wb = openpyxl.load_workbook(file_name)
            
            # 如果指定了工作表，只处理该工作表
            if order_sheet_name:
                if order_sheet_name not in wb.sheetnames:
                    raise ValueError(f"订单文件中不存在工作表: {order_sheet_name}")
                logger(f"仅处理工作表: '{order_sheet_name}'")
            else:
                logger(f"找到 {len(wb.sheetnames)} 个工作表")
        except Exception as e:
            raise Exception(f"加载订单文件失败: {e}")
    
    stats = {
        'sheets_processed': 0,
        'rows_filled': 0
    }
    
    total_sheets_processed = 0
    total_rows_filled = 0
    
    # 4. 遍历每个worksheet填充数据
    # 如果指定了工作表，只处理该工作表；否则处理所有工作表
    worksheets_to_process = [wb[order_sheet_name]] if order_sheet_name else wb.worksheets
    
    for target_ws in worksheets_to_process:
        logger(f"  > 正在检查工作表: {target_ws.title}")
        
        target_idx = {}
        # 获取表头 (row 1)
        target_headers = [c.value for c in target_ws[1]]
        
        if not target_col_map:
            col_map_config = {'sku': 'SKU', 'qty': '数量', 'l': '长', 'w': '宽', 'h': '高', 'wt': '单件重量'}
        else:
            col_map_config = target_col_map

        def get_col_index(config_val):
            # 1. 优先尝试当做表头名称查找
            if config_val in target_headers:
                return target_headers.index(config_val) + 1
            
            # 2. 尝试当做列字母 (A, B, AA...)
            if config_val and isinstance(config_val, str) and config_val.isalpha() and len(config_val) <= 3:
                 try: return column_index_from_string(config_val.upper())
                 except: pass
            
            return None

        missing_cols = []
        for role in ['sku', 'qty', 'l', 'w', 'h', 'wt']:
            val = col_map_config.get(role)
            idx = get_col_index(val)
            if idx:
                target_idx[role] = idx
            else:
                missing_cols.append(f"{role}({val})")

        if missing_cols:
            logger(f"    ...警告：无法定位以下列: {', '.join(missing_cols)}。跳过此表。")
            continue
            
        logger(f"    ...列定位成功，开始处理...")
        logger(f"    Debug: target_idx={target_idx}, max_row={target_ws.max_row}")
        count = 0
        
        # 缓存列索引
        c_sku = target_idx['sku']
        c_qty = target_idx['qty']
        c_l = target_idx['l']
        c_w = target_idx['w']
        c_h = target_idx['h']
        c_wt = target_idx['wt']
        
        # 批量读取数据优化？ openpyxl 写操作通常是瓶颈，读操作还行
        # 这里逐行处理以便同时读写
        for i in range(2, target_ws.max_row + 1):
            sku_val = target_ws.cell(row=i, column=c_sku).value
            if not sku_val: continue
            
            # 读取数量（如果忽略数量，固定为1）
            qty_val = 1
            if not ignore_qty:
                try:
                    q_v = target_ws.cell(row=i, column=c_qty).value
                    if q_v:
                        qty_val = float(q_v)
                        if qty_val <= 0: qty_val = 1
                except:
                    pass
            
            # 处理逻辑
            bundle_list = parse_sku_bundle(sku_val, logger)
            
            final_l = final_w = final_h = final_wt = 0
            valid = False
            err_msg = None
            
            if bundle_list:
                tot_vol = 0; tot_wt = 0; all_found = True
                for sub_sku, sub_qty in bundle_list:
                    data = sku_database.get(sub_sku)
                    if data:
                        tot_vol += data["volume"] * sub_qty
                        tot_wt += data["weight"] * sub_qty
                    else:
                        all_found = False; break
                
                if all_found:
                    final_l = 10; final_w = 8
                    final_h = tot_vol / 80 if tot_vol > 0 else 0
                    final_wt = tot_wt
                    valid = True
                else:
                    err_msg = "组合SKU错误"
            else:
                data = sku_database.get(sku_val)
                if data:
                    new_h = data["height"] * qty_val
                    new_wt = data["weight"] * qty_val
                    dims = sorted([data["length"], data["width"], new_h], reverse=True)
                    final_l, final_w, final_h = dims
                    final_wt = new_wt
                    valid = True
                else:
                    err_msg = "SKU未找到"
            
            if valid:
                target_ws.cell(row=i, column=c_l).value = final_l
                target_ws.cell(row=i, column=c_w).value = final_w
                target_ws.cell(row=i, column=c_h).value = final_h
                target_ws.cell(row=i, column=c_wt).value = final_wt
                count += 1
            elif err_msg:
                # 记录错误原因但不中断
                # target_ws.cell(row=i, column=c_l).value = err_msg
                pass

        logger(f"    处理完成: 成功填充 {count} 行")
        if count > 0:
            total_sheets_processed += 1
            total_rows_filled += count

    try:
        logger(f"正在保存文件: {output_file_name}...")
        wb.save(output_file_name)
        wb.close()
        
        # 如果是从.xls转换的，额外提示用户
        if is_xls_order:
            logger(f"  注意: 由于.xls格式限制，结果已保存为新文件: {os.path.basename(output_file_name)}")
        
        return {'sheets_processed': total_sheets_processed, 'rows_filled': total_rows_filled}
    except PermissionError:
        raise PermissionError(f"无法保存文件 '{output_file_name}'。请检查文件是否已在 Excel/WPS 中打开。")
    except Exception as e:
        raise Exception(f"保存文件时发生未知错误: {e}")
