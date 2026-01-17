from excel_toolkit.excel_lite import ExcelReader, ExcelWriter
import os
from typing import Dict, Callable, Optional

# 导入openpyxl用于写入Excel文件
try:
    import openpyxl
    from openpyxl.styles import PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    openpyxl = None
    PatternFill = None


def _check_file_locked(file_path: str) -> bool:
    """检查文件是否被其他程序锁定（如Excel/WPS）
    
    Returns:
        True 如果文件被锁定，False 如果可以访问
    """
    try:
        with open(file_path, 'a'):  # 尝试追加模式打开
            pass
        return False
    except (IOError, PermissionError):
        return True


def process_insert_rows(
    file_x_path: str, 
    sheet_x_name: str, 
    file_y_path: str, 
    sheet_y_name: str, 
    logger: Callable[[str], None] = print
) -> Dict[str, int]:
    """
    对比两个表格，将X中有但Y中没有的行插入到Y中。
    
    Returns:
        Dict with keys: 'inserted_rows', 'missing_count'
        
    Raises:
        FileNotFoundError, PermissionError, Exception
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("需要openpyxl库来处理Excel文件，请安装: pip install openpyxl")
    
    blue_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
    
    if not os.path.exists(file_x_path):
        raise FileNotFoundError(f"表格X未找到: {file_x_path}")
    if not os.path.exists(file_y_path):
        raise FileNotFoundError(f"表格Y未找到: {file_y_path}")
    
    # 先检查文件是否被锁定（尤其是表格Y需要写入）
    if _check_file_locked(file_y_path):
        raise PermissionError(f"表格Y被占用，请先关闭 Excel/WPS 后再试: {os.path.basename(file_y_path)}")

    # 1. 读取表格X (源) - 使用 read_only 模式优化性能
    try:
        logger(f"正在加载表格X (源): {file_x_path}...")
        wb_X = ExcelReader(file_x_path, read_only=True, data_only=True)
        if sheet_x_name not in wb_X.sheetnames:
             wb_X.close()
             raise ValueError(f"表格X中未找到子表: '{sheet_x_name}'")
             
        ws_X = wb_X[sheet_x_name]
        logger(f"从表格X的 '{sheet_x_name}' 子表中读取数据...")
        
        x_pairs = set()
        # iter_rows with values_only=True is faster
        for row in ws_X.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
            pid_raw = row[0]
            iid_raw = row[1]
            if pid_raw is None or pid_raw == "":
                continue
            if iid_raw is None or iid_raw == "":
                continue
            x_pairs.add((str(pid_raw), str(iid_raw)))
            
        wb_X.close()
        logger(f"从表格X中读取了 {len(x_pairs)} 个唯一的“商品号-ID”组合。")
        
    except Exception as e:
        raise Exception(f"读取表格X失败: {e}")

    # 2. 读取表格Y (目标) - 需要写入，使用openpyxl
    try:
        logger(f"正在加载表格Y (目标): {file_y_path}...")
        wb_Y = openpyxl.load_workbook(file_y_path)
        if sheet_y_name not in wb_Y.sheetnames:
             wb_Y.close()
             raise ValueError(f"表格Y中未找到子表: '{sheet_y_name}'")
             
        ws_Y = wb_Y[sheet_y_name]
        logger(f"从表格Y的 '{sheet_y_name}' 子表中读取数据...")
        
        y_pairs = set()
        last_row_map = {}
        
        # 这里不能用 values_only=True 因为我们需要 row index
        # 但是我们可以优化：只读取前两列
        for r in range(2, ws_Y.max_row + 1):
            pid_raw = ws_Y.cell(row=r, column=1).value
            iid_raw = ws_Y.cell(row=r, column=2).value
            
            if pid_raw is None or pid_raw == "":
                continue
            if iid_raw is None or iid_raw == "":
                continue
                
            pid = str(pid_raw)
            iid = str(iid_raw)
            y_pairs.add((pid, iid))
            last_row_map[pid] = r
            
        logger(f"从表格Y中读取了 {len(y_pairs)} 个“商品号-ID”组合。")

        missing_pairs = x_pairs - y_pairs
        if not missing_pairs:
            wb_Y.close()
            return {'inserted_rows': 0, 'missing_count': 0}
            
        logger(f"找到了 {len(missing_pairs)} 个缺失的行。")

        missing_grouped = {}
        for pair in missing_pairs:
            pid = pair[0]
            if pid not in missing_grouped:
                missing_grouped[pid] = []
            missing_grouped[pid].append(pair)

        def get_row_key(pid): return last_row_map.get(pid, 0)
        sorted_pids_to_insert = sorted(missing_grouped.keys(), key=get_row_key, reverse=True)

        logger("开始在表格Y中插入行...")
        total_inserted = 0
        
        for pid in sorted_pids_to_insert:
            insertion_row = last_row_map.get(pid)
            if not insertion_row:
                logger(f"警告：商品号 '{pid}' 在表格Y中不存在，无法为其插入新行。")
                continue
                
            pairs_to_add = missing_grouped[pid]
            # 倒序插入，保持顺序
            for pair in reversed(pairs_to_add):
                ws_Y.insert_rows(insertion_row + 1)
                c1 = ws_Y.cell(row=insertion_row + 1, column=1)
                c2 = ws_Y.cell(row=insertion_row + 1, column=2)
                c1.value = pair[0]
                c2.value = pair[1]
                c1.fill = blue_fill
                c2.fill = blue_fill
                total_inserted += 1
                
        if total_inserted == 0:
            wb_Y.close()
            return {'inserted_rows': 0, 'missing_count': len(missing_pairs)}

        wb_Y.save(file_y_path)
        wb_Y.close()
        return {'inserted_rows': total_inserted, 'missing_count': len(missing_pairs)}
        
    except PermissionError:
        raise PermissionError(f"无法保存文件 '{file_y_path}'。请检查文件是否已在 Excel/WPS 中打开。")
    except Exception as e:
        raise Exception(f"处理表格Y时发生错误: {e}")
