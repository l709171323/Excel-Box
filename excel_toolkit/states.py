from typing import Callable, Dict, Optional, Union
import os
try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    from excel_toolkit.excel_lite import column_index_from_string
    openpyxl = None

from excel_toolkit.exceptions import (
    FileLockedError,
    FileNotFoundError as CustomFileNotFoundError,
    InvalidColumnError,
    SheetNotFoundError
)
from excel_toolkit.error_handler import log_error, handle_file_error


STATE_MAP = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR", "california": "CA",
    "colorado": "CO", "connecticut": "CT", "delaware": "DE", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS", "missouri": "MO",
    "montana": "MT", "nebraska": "NE", "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC", "north dakota": "ND",
    "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI",
    "south carolina": "SC", "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY", "puerto rico": "PR", "guam": "GU",
    "u.s. virgin islands": "VI", "us virgin islands": "VI",
    "district of columbia": "DC", "washington dc": "DC", "washington d.c.": "DC"
}


def get_state_abbreviation(state_name: str) -> Optional[str]:
    """将州全名转换为缩写
    
    Args:
        state_name: 州的全名（如 'California'）
    
    Returns:
        州的缩写（如 'CA'），未找到返回 None
    """
    if not state_name:
        return None
    lookup_name = str(state_name).strip().lower()
    return STATE_MAP.get(lookup_name)


def process_states(
    file_name: str, 
    sheet_name: str, 
    column_letter: str, 
    logger: Callable[[str], None] = print
) -> Dict[str, int]:
    """处理Excel文件中的州名转换
    
    Args:
        file_name: Excel文件路径
        sheet_name: 工作表名称
        column_letter: 要处理的列号（如 'A', 'B'）
        logger: 日志输出函数
    
    Returns:
        包含统计信息的字典 {'success': int, 'failed': int, 'total': int}
        
    Raises:
        CustomFileNotFoundError: 文件不存在
        InvalidColumnError: 列号无效
        FileLockedError: 文件被占用
        SheetNotFoundError: 工作表不存在
    """
    # 验证文件存在
    if not os.path.exists(file_name):
        raise CustomFileNotFoundError(file_name)

    # 验证openpyxl可用
    if openpyxl is None:
        raise Exception("需要安装 openpyxl 库才能修改 Excel 文件")

    # 加载工作簿
    try:
        wb = openpyxl.load_workbook(file_name)
    except PermissionError as e:
        handle_file_error(file_name, e)
    except Exception as e:
        log_error(e, f"加载文件: {file_name}")
        raise Exception(f"加载文件失败: {e}")

    # 验证工作表存在
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SheetNotFoundError(sheet_name, wb.sheetnames)
    
    ws = wb[sheet_name]
    
    # 验证列号
    try:
        col_index = column_index_from_string(column_letter)
    except ValueError:
        wb.close()
        raise InvalidColumnError(column_letter)

    stats = {'success': 0, 'failed': 0, 'total': 0}
    
    logger(f"开始处理工作表 '{sheet_name}' 的 {column_letter} 列...")
    
    # 获取最大行数，避免处理空行
    max_row = ws.max_row
    
    for i in range(2, max_row + 1):
        cell = ws.cell(row=i, column=col_index)
        full_name = cell.value
        
        if full_name:
            stats['total'] += 1
            abbr = get_state_abbreviation(full_name)
            if abbr:
                cell.value = abbr
                stats['success'] += 1
            else:
                stats['failed'] += 1
                # 保持原值，不修改
                # logger(f"  行 {i}: 未找到州名 '{full_name}'，保持原值") 

    # 保存文件
    try:
        wb.save(file_name)
        wb.close()
        logger(f"✅ 文件已成功保存")
        return stats
    except PermissionError as e:
        wb.close()
        log_error(e, f"保存文件: {file_name}")
        raise FileLockedError(os.path.basename(file_name))
    except Exception as e:
        wb.close()
        log_error(e, f"保存文件: {file_name}")
        raise Exception(f"保存文件时发生错误: {e}")

