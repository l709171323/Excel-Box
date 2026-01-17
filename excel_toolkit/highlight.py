from typing import Callable, Dict
from excel_toolkit.excel_lite import column_index_from_string
import os

# 导入openpyxl用于读写Excel文件（需要支持PatternFill格式）
try:
    import openpyxl
    from openpyxl.styles import PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    openpyxl = None
    PatternFill = None

# 常量定义
YELLOW_HIGHLIGHT = "FFFFFF00"
ORANGE_HIGHLIGHT = "FFFFC000"


def _check_file_locked(file_path: str) -> bool:
    """检查文件是否被其他程序锁定（如Excel/WPS）
    
    Returns:
        True 如果文件被锁定，False 如果可以访问
    """
    try:
        with open(file_path, 'a'):  # 尝试追加模式打开
            pass
        return False
    except (IOError, PermissionError):
        return True


def highlight_duplicates(
    file_name: str, 
    column_letter: str, 
    logger: Callable[[str], None] = print,
    sheet_name: str = None
) -> Dict[str, int]:
    """高亮Excel文件中指定列的重复项
    
    Args:
        file_name: Excel文件路径
        column_letter: 要检查的列号（如 'A', 'B'）
        logger: 日志输出函数
        sheet_name: 工作表名称，如果为None则处理所有工作表
    
    Returns:
        Dict with keys: 'sheets_processed', 'cells_highlighted'
        
    Raises:
        FileNotFoundError, ValueError, PermissionError, Exception
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("需要openpyxl库来处理Excel文件，请安装: pip install openpyxl")
    
    yellow_fill = PatternFill(start_color=YELLOW_HIGHLIGHT, end_color=YELLOW_HIGHLIGHT, fill_type="solid")
    orange_fill = PatternFill(start_color=ORANGE_HIGHLIGHT, end_color=ORANGE_HIGHLIGHT, fill_type="solid")
    colors = [yellow_fill, orange_fill]
    
    logger(f"准备加载文件: {os.path.basename(file_name)}")
    
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"文件未找到: {file_name}")

    # 先检查文件是否被锁定
    if _check_file_locked(file_name):
        raise PermissionError(f"文件被占用，请先关闭 Excel/WPS 后再试: {os.path.basename(file_name)}")

    try:
        logger("正在加载Excel文件，请稍候...")
        wb = openpyxl.load_workbook(file_name)
        logger("文件加载完成。")
    except Exception as e:
        raise Exception(f"加载文件失败: {e}")
        
    try:
        col_index = column_index_from_string(column_letter)
    except ValueError:
        wb.close()
        raise ValueError(f"列号 '{column_letter}' 无效")
        
    total_sheets_processed = 0
    total_cells_highlighted = 0
    
    # 如果指定了工作表，只处理该工作表；否则处理所有工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        worksheets_to_process = [wb[sheet_name]]
        logger(f"开始处理工作表 '{sheet_name}'，高亮 {column_letter} 列...")
    else:
        worksheets_to_process = wb.worksheets
        logger(f"开始遍历所有工作表，高亮 {column_letter} 列...")
    
    for ws in worksheets_to_process:
        if ws.max_row <= 1:
            continue
            
        logger(f"  > 正在处理工作表: {ws.title}")
        value_rows = {}
        
        # 第一遍遍历：收集数据
        for row in ws.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
            cell = row[0]
            if cell.value is None or cell.value == "":
                continue
            value = str(cell.value)
            if value not in value_rows:
                value_rows[value] = []
            value_rows[value].append(cell.row)
            
        color_index = 0
        sheet_cells_count = 0
        
        # 第二遍遍历：高亮重复项
        for value, rows in value_rows.items():
            if len(rows) > 1:
                current_color = colors[color_index]
                for row_num in rows:
                    ws.cell(row=row_num, column=col_index).fill = current_color
                    sheet_cells_count += 1
                color_index = 1 - color_index
                
        if sheet_cells_count > 0:
            total_sheets_processed += 1
            total_cells_highlighted += sheet_cells_count
            
    try:
        wb.save(file_name)
        wb.close()
        return {
            'sheets_processed': total_sheets_processed,
            'cells_highlighted': total_cells_highlighted
        }
    except PermissionError:
        raise PermissionError(f"无法保存文件 '{file_name}'。请检查文件是否已在 Excel/WPS 中打开。")
    except Exception as e:
        raise Exception(f"保存文件时发生未知错误: {e}")