"""
批量删除表格列功能

用于从Excel文件中批量删除指定的列（如D、E列）。
"""

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os


def parse_column_input(columns_str: str) -> list:
    """
    解析用户输入的列标识字符串
    
    支持的格式：
    - "D,E" 或 "D, E" - 逗号分隔
    - "D E F" - 空格分隔
    - "D-F" - 范围格式（删除D到F列）
    
    返回列标识列表（已排序，从大到小，以便从后往前删除）
    """
    columns_str = columns_str.strip().upper()
    if not columns_str:
        return []
    
    result = []
    
    # 处理范围格式 (如 "D-F")
    if '-' in columns_str and ',' not in columns_str and ' ' not in columns_str:
        parts = columns_str.split('-')
        if len(parts) == 2:
            start_col = parts[0].strip()
            end_col = parts[1].strip()
            if start_col.isalpha() and end_col.isalpha():
                start_idx = column_index_from_string(start_col)
                end_idx = column_index_from_string(end_col)
                if start_idx > end_idx:
                    start_idx, end_idx = end_idx, start_idx
                for i in range(start_idx, end_idx + 1):
                    from openpyxl.utils import get_column_letter
                    result.append(get_column_letter(i))
                # 从大到小排序（便于从后往前删除）
                result.sort(key=lambda x: column_index_from_string(x), reverse=True)
                return result
    
    # 处理逗号或空格分隔的格式
    if ',' in columns_str:
        parts = columns_str.split(',')
    else:
        parts = columns_str.split()
    
    for part in parts:
        col = part.strip()
        if col and col.isalpha():
            result.append(col)
    
    # 去重并从大到小排序（便于从后往前删除）
    result = list(set(result))
    result.sort(key=lambda x: column_index_from_string(x), reverse=True)
    
    return result


def delete_columns(file_path: str, columns: list, logger=None, sheet_name: str = None) -> dict:
    """
    从Excel文件中删除指定的列
    
    Args:
        file_path: Excel文件路径
        columns: 要删除的列列表 (如 ['D', 'E'])，应该已经从大到小排序
        logger: 日志回调函数
        sheet_name: 可选，指定工作表名称（为空则处理所有工作表）
    
    Returns:
        dict: 统计信息 {'sheets_processed': int, 'columns_deleted': int}
    """
    def log(msg):
        if logger:
            logger(msg)
        else:
            print(msg)
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    if not columns:
        raise ValueError("未指定要删除的列")
    
    # 验证列标识
    for col in columns:
        if not col.isalpha():
            raise ValueError(f"无效的列标识: {col}")
        try:
            column_index_from_string(col)
        except ValueError:
            raise ValueError(f"无效的列标识: {col}")
    
    log(f"正在加载文件: {os.path.basename(file_path)}")
    
    # 加载工作簿
    wb = load_workbook(file_path)
    
    # 确定要处理的工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        sheets_to_process = [sheet_name]
    else:
        sheets_to_process = wb.sheetnames
    
    stats = {
        'sheets_processed': 0,
        'columns_deleted': 0
    }
    
    # 确保columns是从大到小排序的（从后往前删除，避免列索引变化问题）
    sorted_columns = sorted(columns, key=lambda x: column_index_from_string(x), reverse=True)
    
    for sname in sheets_to_process:
        ws = wb[sname]
        log(f"  处理工作表: {sname}")
        
        deleted_in_sheet = 0
        for col in sorted_columns:
            try:
                col_idx = column_index_from_string(col)
                if col_idx <= ws.max_column:
                    ws.delete_cols(col_idx)
                    deleted_in_sheet += 1
                    log(f"    ✓ 已删除列 {col}")
                else:
                    log(f"    ⚠ 列 {col} 超出工作表范围，跳过")
            except Exception as e:
                log(f"    ✗ 删除列 {col} 失败: {e}")
        
        stats['columns_deleted'] += deleted_in_sheet
        stats['sheets_processed'] += 1
    
    # 保存文件
    log(f"正在保存文件...")
    wb.save(file_path)
    wb.close()
    
    log(f"✅ 保存完成!")
    
    return stats
